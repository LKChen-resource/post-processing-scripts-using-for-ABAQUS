from odbAccess import *
import openpyxl as op


filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\TA\OutPutResults_TA.xlsx'
wb=op.load_workbook(filename)
sh=wb["Sheet1"]


for ii in range(1,2):
    print('begin'+str(ii))
    ##openODB
    ModelName=str(sh.cell(row=ii,column=1).value)
    JobName=ModelName.replace('-','_')
    odb=openOdb(path=JobName+'.odb')
    D=sh.cell(row=ii,column=5).value
    print(D)
    As=3.14159*D**2/4
    H=sh.cell(row=ii,column=6).value
    ##select step
    step2=odb.steps.values()[-1]

    ##put every frame to list 'Frame'
    Frame=[]
    RFvalues=[]
    DispValue=[]
    for i in range(0,100):
        # print(i)
        Frame.append(odb.steps[step2.name].frames[i])
        # print(Frame)
        ##Reaction Force at node2
        reactionForce=Frame[i].fieldOutputs['RF']
        rfValues=reactionForce.values
        list1=[]#cylindal node set
        # for l in range(1,450):
        #     list1.append(l)
        print(rfValues)

        for r in rfValues:
           if r.nodeLabel == 1 :#the assemble node 1 would repeat another node 1,so use the another node 
                if r.data[1]<-1:
                 RFvalues.append(r.data[1])
                # print(r.nodeLabel)
        # print(RFvalues)
        
        ##Axial Starin
        displacement=Frame[i].fieldOutputs['U']
        dispvalue=displacement.values
        for d in dispvalue:
            
            if d.nodeLabel==2:#the assemble node 1 would repeat another node 1,so use the another node 
                DispValue.append(d.data[1])
    # print(RFvalues)
    # print(DispValue)

    # filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\TA\OutPutResults_TA.xlsx'
    # wb=op.load_workbook(filename)
    # sh=wb["U-RF"]

    # for i in range(1,len(DispValue)+1):
    #     # sh.cell(row=i+1,column=1).value=float(10)
    #     sh.cell(row=i+1,column=2*ii).value=float(-DispValue[i-1]/H)
    #     sh.cell(row=i+1,column=2*ii+1).value=float(-RFvalues[i-1]/As)


    # wb.save(filename)



    





