from odbAccess import *
import openpyxl as op


filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\CFST_Parameter_Analyse.xlsx'
wb=op.load_workbook(filename)
sh=wb["Sheet1"]
sh2=wb["Sheet3"]


for ii in range(28,29):
    print('begin'+str(ii))
    ##openODB
    ModelName=str(sh.cell(row=ii,column=0).value)
    JobName=ModelName.replace('-','_')
    odb=openOdb(path=JobName+'.odb')
    D=sh.cell(row=ii,column=3).value
    t=sh.cell(row=ii,column=4).value
    Ac=3.14159*(D-2*t)**2/4
    As=3.14159*D**2/4-3.14159*(D-2*t)**2/4
    H=sh.cell(row=ii,column=6).value
    ##select step
    step2=odb.steps.values()[-1]

    ##put every frame to list 'Frame'
    Frame=[]
    ConcreteRFvalues=[]
    SteelRFvalues=[]
    DispValue=[]

    for XPpoint in ["ASSEMBLY_CONSTRAINT-3_REFERENCE_POINT","ASSEMBLY_CONSTRAINT-4_REFERENCE_POINT"]:
        RFvalues=[]
        for i in range(0,79):
            # print(i)
            Frame.append(odb.steps[step2.name].frames[i])
            # print(Frame)
            ##Reaction Force at node2
            reactionForce=Frame[i].fieldOutputs['RF']
            XPnode1=odb.rootAssembly.nodeSets[XPpoint]
            XPforce=reactionForce.getSubset(region=XPnode1)
            rfValues=XPforce.values
            list1=[]#cylindal node set
            # for l in range(1,450):
            #     list1.append(l)
            # print(rfValues)
            RFvalues.append(XPforce.values[0].data[2])
            if XPpoint=="ASSEMBLY_CONSTRAINT-3_REFERENCE_POINT":
                displacement=Frame[i].fieldOutputs['U']
                dispvalue=displacement.values
                XPdisp=displacement.getSubset(region=XPnode1)
                DispValue.append(XPdisp.values[0].data[2])
            
        if XPpoint=="ASSEMBLY_CONSTRAINT-3_REFERENCE_POINT":
            SteelRFvalues=RFvalues
            
            #print(ConcreteRFvalues)
        else:
            ConcreteRFvalues=RFvalues


#rint(len(DispValue),len(ConcreteRFvalues),len(SteelRFvalues))
    sh2.cell(row=0,column=7*ii-21).value=ModelName
    sh2.cell(row=1,column=7*ii-21).value="axial strain"
    sh2.cell(row=1,column=7*ii-21+1).value="Nc"
    sh2.cell(row=1,column=7*ii-21+2).value="Ns"
    sh2.cell(row=1,column=7*ii-21+3).value="Nu"
    sh2.cell(row=1,column=7*ii-21+4).value="Rc"
    sh2.cell(row=1,column=7*ii-21+5).value="fcc"
    sh2.cell(row=1,column=7*ii-21+6).value="sigmaz"
    # D=float(sh.cell(row=ii,column=3).value)
    # print(D)
    # t=float(sh.cell(row=ii,column=4).value)
    for i in range(0,len(DispValue)):
        Nc=float(-ConcreteRFvalues[i]/1000)
        Ns=float(-SteelRFvalues[i]/1000)
        # Ac=3.14159*(D-2*t)**2/4
        # As=3.14159*D**2/4-3.14159*(D-2*t)**2/4
        
        sh2.cell(row=i+2,column=7*ii-21).value=float(-DispValue[i]/H)
        sh2.cell(row=i+2,column=7*ii-21+1).value=float(-ConcreteRFvalues[i]/1000)
        sh2.cell(row=i+2,column=7*ii-21+2).value=float(-SteelRFvalues[i]/1000)
        sh2.cell(row=i+2,column=7*ii-21+3).value=Nc+Ns#Nu
        if i==0:
            sh2.cell(row=i+2,column=7*ii-21+4).value=0.5#Rc
        else:
            sh2.cell(row=i+2,column=7*ii-21+4).value=Nc/(Nc+Ns)#Rc
        sh2.cell(row=i+2,column=7*ii-21+5).value=Nc*1000/Ac#fcc
        sh2.cell(row=i+2,column=7*ii-21+6).value=Ns*1000/As#sigmaz
        # sh2.cell(row=i+2,column=4*ii-8+3).value=float(-ConcreteRFvalues[i]/(-SteelRFvalues[i]-ConcreteRFvalues[i]))


wb.save(filename)
print("End----")



    





