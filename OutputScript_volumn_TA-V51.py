from odbAccess import *
from textRepr  import *
import openpyxl as op

for ii in range(0,1):  # number of jods
    print('begin'+str(ii))
    ##openODB
    #odb=openOdb(path='0425-'+'%02d'%(ii+1)+'.odb')
    odb=openOdb(path='0425-17.odb')
    ##select step
    step2=odb.steps.values()[-1]

    ##put every frame to list 'Frame'
    Frame=[]
    sumVolumn=[]
    RFvalues=[]
    DispValue=[]
    LateralStrain=[]
    AxialStrain=[]
    for i in range(0,100):
        #print(i)
        Frame.append(odb.steps[step2.name].frames[i])
        #print(Frame)

        ##volumn
        volumn=Frame[i].fieldOutputs['EVOL']
        volumnValues=volumn.values
        sumVolumnFrame=0
        for v in volumnValues:
            # print(v.data)
            sumVolumnFrame=sumVolumnFrame +v.data
        sumVolumn.append(sumVolumnFrame)

        XP_1=odb.rootAssembly.nodeSets['ASSEMBLY_CONSTRAINT-1_REFERENCE_POINT']
        ##Axial Starin
        displacement=Frame[i].fieldOutputs['U']
        XP_1displace=displacement.getSubset(region=XP_1)
        XP_1displaceValue=XP_1displace.values
        for d in XP_1displaceValue:
            #DispValue.append(d.data[1]/100)
            AxialStrain.append(d.data[1]/100)
    # print(DispValue)

        ##Reaction Force at XP-1
        reactionForce=Frame[i].fieldOutputs['RF']
        XP_1RF=reactionForce.getSubset(region=XP_1)
        XP_1RFValue=XP_1RF.values
        for r in XP_1RFValue:
            RFvalues.append(r.data[1])
        # print(RFvalues)
        ####-------reference-----------
        # XP_1=odb.rootAssembly.nodeSets['ASSEMBLY_CONSTRAINT-1_REFERENCE_POINT']
        # XP_1displace=displacement.getSubset(region=XP_1)
        # XP_1displaceValue=XP_1displace.values
        # for v in XP_1displaceValue:
        #     print v.nodeLabel,v.data[1]
        # XP_1RF=reactionForce.getSubset(region=XP_1)
        # XP_1RFValue=XP_1RF.values
        # for v in XP_1RFValue:
        #     print v.nodeLabel,v.data[1]
        ####----------end----------------

    #print(sumVolumn)##volumn in each step
    ##caculate volumetricStrain
    
    # volumnStrain=[0]
    # for i in range(len(sumVolumn)-1):
    #     volumnStrain.append((sumVolumn[i+1]-sumVolumn[0])/sumVolumn[0])
    # print(volumnStrain)
    # print(len(DispValue),len(volumnStrain))
    
    ##caculat axial strain

    ##caculate lateral Strain
    print(sumVolumn[0])
    lateralStrain=[0]
    for i in range(len(sumVolumn)-1):
        lateralStrain.append(((sumVolumn[i+1]-sumVolumn[0])/sumVolumn[0]-AxialStrain[i+1])/2)
    # print(lateralStrain)
    # print(AxialStrain)
    # print(RFvalues)
    # print(len(DispValue),len(volumnStrain))

    
   

    filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\TA\OutPutResults_TA.xlsx'
    wb=op.load_workbook(filename)
    sh=wb["eps3-eps1"]
    for i in range(1,len(AxialStrain)+1):
        # sh.cell(row=i+1,column=1).value=float(10)
        #sh.cell(row=i+3,column=3).value=float(AxialStrain[i-1])
        #sh.cell(row=i+3,column=4*ii+1).value=float(RFvalues[i-1]/-10000)
        sh.cell(row=i+1,column=3).value=float(lateralStrain[i-1])
        sh.cell(row=i+1,column=4).value=float(RFvalues[i-1]/-10000)

    wb.save(filename)
print('---END---')
