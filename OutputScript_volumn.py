from odbAccess import *
import openpyxl as op

for ii in range(9):
    print('begin'+str(ii))
    ##openODB
    odb=openOdb(path='220114_'+'%02d'%(ii+3)+'.odb')
    ##select step
    step2=odb.steps.values()[-1]

    ##put every frame to list 'Frame'
    Frame=[]
    sumVolumn=[]
    DispValue=[]
    for i in range(0,100):
        # print(i)
        Frame.append(odb.steps[step2.name].frames[i])
        # print(Frame)
        volumn=Frame[i].fieldOutputs['EVOL']
        volumnValues=volumn.values
        sumVolumnFrame=0
        for v in volumnValues:
            # print(v.data)
            sumVolumnFrame=sumVolumnFrame +v.data
        sumVolumn.append(sumVolumnFrame)
        ##Axial Starin
        displacement=Frame[i].fieldOutputs['U']
        dispvalue=displacement.values
        for d in dispvalue:
            
            if d.nodeLabel==2:#the assemble node 1 would repeat another node 1,so use the another node 
                DispValue.append(d.data[1])
    # print(DispValue)

    #print(sumVolumn)##volumn in each step
    ##caculate volumetricStrain
    volumnStrain=[0]
    for i in range(len(sumVolumn)-1):
        volumnStrain.append((sumVolumn[i+1]-sumVolumn[0])/sumVolumn[0])
    # print(volumnStrain)
    # print(len(DispValue),len(volumnStrain))
    ##Axial Starin

    filename=r'C:\Users\LKChen\Desktop\HarryHardingCruve.xlsx'
    wb=op.load_workbook(filename)
    sh=wb["Sheet3"]
    for i in range(1,len(DispValue)+1):
        # sh.cell(row=i+1,column=1).value=float(10)
        sh.cell(row=i+1,column=2*ii).value=float(-DispValue[i-1]/108)
        sh.cell(row=i+1,column=2*ii+1).value=float(volumnStrain[i-1])


    wb.save(filename)



    





