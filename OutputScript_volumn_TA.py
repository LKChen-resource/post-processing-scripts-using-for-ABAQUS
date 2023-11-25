from odbAccess import *
import openpyxl as op
#import matplotlib.pyplot as plt

filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\Uniaxial_Biaxial_Triaxial_Test.xlsx'
wb=op.load_workbook(filename)
sh=wb["ABAQUSDataBase"]

for ii in range(54,55):
    
    ##openODB
    ModelName=str(sh.cell(row=ii,column=1).value)
    JobName=ModelName
    JobName=ModelName.replace('-','_')
    print('begin'+JobName)
    odb=openOdb(path=JobName+'.odb')
    #H=sh.cell(row=ii,column=6).value
    #D=sh.cell(row=ii,column=5).value
    #As=3.14159*D**2/4
    #As=10000
    ##select step
    step2=odb.steps.values()[-1]

    ##put every frame to list 'Frame'
    Frame=[]
    sumVolumn=[]
    AxialStrain=[]
    displaceValue=[]
    RFvalues=[]
    for i in range(0,101):
        # print(i)
        Frame.append(odb.steps[step2.name].frames[i])
        # print(Frame)
        volumn=Frame[i].fieldOutputs['EVOL']
        volumnValues=volumn.values
        sumVolumnFrame=0
        for v in volumnValues:
            # print(v.data)
            sumVolumnFrame=sumVolumnFrame +v.data
        sumVolumn.append(sumVolumnFrame)
        ##Axial Starin
        displacement=Frame[i].fieldOutputs['U']
        RP_11=odb.rootAssembly.nodeSets['ASSEMBLY_CONSTRAINT-1_REFERENCE_POINT']
        U1=displacement.getSubset(region=RP_11)
        displaceValue.append(U1.values[0].data[2])
        AxialStrain.append(displaceValue[i]/100.)############

        ##Reaction Force at node2
        reactionForce=Frame[i].fieldOutputs['RF']
        RP_1=odb.rootAssembly.nodeSets['ASSEMBLY_CONSTRAINT-1_REFERENCE_POINT']
        RF1=reactionForce.getSubset(region=RP_1)
        #prettyPrint(k,2)
        RFvalues.append(RF1.values[0].data[2])################
   
    # print(DispValue)

    #print(sumVolumn)##volumn in each step
    ##caculate volumetricStrain
    # print(sumVolumn[0])
    lateralStrain=[0]
    volumnStrain=[0]
    for i in range(len(sumVolumn)-1):
        lateralStrain.append(((sumVolumn[i+1]-sumVolumn[0])/sumVolumn[0]-AxialStrain[i+1])/2.)
        volumnStrain.append((sumVolumn[i+1]-sumVolumn[0])/sumVolumn[0])
    # print(volumnStrain)
    # print(len(DispValue),len(volumnStrain))
    ##Axial Starin

    filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\TA\OutPutResults_TA_CDP.xlsx'
    wb=op.load_workbook(filename)


    sh1=wb["eps3-eps1"]
    sh1.cell(row=0,column=2*ii).value=ModelName+'eps-3'
    sh1.cell(row=0,column=2*ii+1).value=ModelName+'eps-1'
    # sh1.cell(row=0,column=2*ii).value=ModelName+'eps-1'
    # sh1.cell(row=0,column=2*ii+1).value=ModelName+'sigma-3'
    for j in range(1,len(AxialStrain)+1):
        
        # sh1.cell(row=j+1,column=2*ii).value=float(-AxialStrain[j-1])
        #sh1.cell(row=j+1,column=2*ii+1).value=float(lateralStrain[j-1])
        #sh1.cell(row=j+1,column=2*ii+1).value=float(volumnStrain[j-1])
        sh1.cell(row=j+1,column=2*ii).value=float(-lateralStrain[j-1])
        sh1.cell(row=j+1,column=2*ii+1).value=float(-RFvalues[j-1]/10000.)
        # sh1.cell(row=j+1,column=0).value=float(-lateralStrain[j-1])
        # sh1.cell(row=j+1,column=1).value=float(-RFvalues[j-1]/10000.)


    wb.save(filename)
    #print(eps3)
print('------End-----')


    





