from odbAccess import *
from textRepr  import *
import openpyxl as op
import visualization



#################
#本代码可以输出指定节点集的接触应力至指定的表格，第一行为各个节点的编号，后续按帧数排列
#代码缺陷在于，输出的接触应力为节点坐标的接触应力，而不是指定自定坐标的接触应力。
#一个是在每个节点出建立局部坐标，然后使用getTransformedField(dtm)
#一个是直接对结果做一个坐标上的转换。

# filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\TA\OutPutResults_TA.xlsx'
# wb=op.load_workbook(filename)
# sh=wb["Sheet1"]


for ii in range(1,2):
    print('begin'+str(ii))
    ##openODB
    # ModelName=str(sh.cell(row=ii,column=1).value)
    # JobName=ModelName.replace('-','_')
    # odb=openOdb(path=JobName+'.odb')

    #-------------------------
    odb=openOdb(path='0705-02.odb')
    # dtm=session.scratchOdbs['0705-02.odb'].rootAssembly.datumCsyses['CSYS-1']
    # session.viewports['Viewport: 1'].odbDisplay.basicOptions.setValues(
    #     transformationType=NODAL)
        #transformationType=USER_SPECIFIED, datumCsys=dtm)
    #------------------------------
    # 
    #     # D=sh.cell(row=ii,column=5).value
    # print(D)
    # As=3.14159*D**2/4
    # H=sh.cell(row=ii,column=6).value
    # ##select step
    step2=odb.steps.values()[-1]

    ##put every frame to list 'Frame'
    Frame=[]
    NLO=[]#node list to ouout
    CNF=[[]]
    framenum=10
    for nn in range(framenum):
        CNF.append([])
    for i in range(0,framenum):
        Frame.append(odb.steps[step2.name].frames[i])
        # print(Frame)
        ##CNORMF at nodeset
        contactForce=Frame[i].fieldOutputs['CNORMF   ASSEMBLY__PICKEDSURF18/ASSEMBLY__PICKEDSURF19']
        
        nodeline=odb.rootAssembly.nodeSets['NodeSet-1']
        #dtm=session.scratchOdbs['0705-02.odb'].rootAssembly.datumCsyses['CSYS-1']
        setvalue=contactForce.getSubset(region=nodeline)
        cfvalues=setvalue.values

        if i==0:
            n=0#the number of node
            for nn in cfvalues:
                NLO.append(nn.nodeLabel)
                n=n+1
           
            
        for j in cfvalues :
            CNF[i].append(j.data[0])

    ## adjust to output style
    CNFO=[[]]
    for i in range(n):
        CNFO.append([])
        for f in range(framenum):
            CNFO[i].append(CNF[f][i])
    print(NLO)
    print(CNFO)

    print('End-----------')    
        
    filename=r'C:\Users\LKChen\Desktop\0701-02.xlsx'
    wb=op.load_workbook(filename)
    sh=wb["Sheet1"]

    for i in range(0,len(NLO)):
        sh.cell(row=0,column=i).value=str(NLO[i])
        for j in range(0,framenum):
            sh.cell(row=j+1,column=i).value=float(CNFO[i][j])

    wb.save(filename)



    





