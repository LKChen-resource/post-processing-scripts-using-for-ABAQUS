from pprint import PrettyPrinter
from odbAccess import *
import openpyxl as op
from textRepr import *


filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\Uniaxial_Biaxial_Triaxial_Test.xlsx'
wb=op.load_workbook(filename)
sh=wb["ABAQUSDataBase"]


for ii in range(0,1):
    print('begin'+str(ii))
    ##openODB
    # ModelName=str(sh.cell(row=ii,column=1).value)
    ModelName='0907-02-0_5'
    print(ModelName)
    # JobName=ModelName.replace('-','_')
    JobName='0907-02-0_5'
    odb=openOdb(path=JobName+'.odb')
    #D=sh.cell(row=ii,column=5).value
    #As=3.14159*D**2/4
    #H=sh.cell(row=ii,column=6).value
    ##select step
    step2=odb.steps.values()[-1]

    ##put every frame to list 'Frame'
    Frame=[]
    RFvalues=[]
    DispValue=[]
    
    
    for i in range(0,101):
        # print(i)
        Frame.append(odb.steps[step2.name].frames[i])
        # print(Frame)
        ##Reaction Force at node2
        reactionForce=Frame[i].fieldOutputs['RF']
        RP_1=odb.rootAssembly.nodeSets['ASSEMBLY_CONSTRAINT-3_REFERENCE_POINT']
        RF1=reactionForce.getSubset(region=RP_1)
        #prettyPrint(k,2)
        RFvalues.append(RF1.values[0].data[2])
    
    
        
        # ##Axial Starin
        displacement=Frame[i].fieldOutputs['U']
        RP_11=odb.rootAssembly.nodeSets['ASSEMBLY_CONSTRAINT-3_REFERENCE_POINT']
        U1=displacement.getSubset(region=RP_11)
        DispValue.append(U1.values[0].data[2])
    # print(RFvalues)
    #print(DispValue)
    #print(As)
    #filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\TA\OutPutResults_TA_CDP.xlsx'
    filename=r'C:\Users\LKChen\Desktop\show.xlsx'
    wb=op.load_workbook(filename)
    sh1=wb["U-RF"]
    sh1.cell(row=0,column=2*ii).value=ModelName+'eps-3'
    sh1.cell(row=0,column=2*ii+1).value=ModelName+'sigma-3'

    for i in range(1,len(DispValue)+1):
        # sh.cell(row=i+1,column=1).value=float(10)
        sh1.cell(row=i+1,column=2*ii).value=float(-DispValue[i-1]/100.)
        sh1.cell(row=i+1,column=2*ii+1).value=float(-RFvalues[i-1]/10000.)
        # sh1.cell(row=i+1,column=0).value=float(-DispValue[i-1]/100.)
        # sh1.cell(row=i+1,column=1).value=float(-RFvalues[i-1]/10000.)


    wb.save(filename)

print('------End-----')


    





